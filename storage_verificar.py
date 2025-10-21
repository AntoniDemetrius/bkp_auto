import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta
import os
import re
from collections import defaultdict
from storage_settings import *
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dateutil.relativedelta import relativedelta
import threading
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

file_cache = {}
cache_lock = threading.Lock()

def retry(func, *args, **kwargs):
    stop_event = kwargs.pop('stop_event', None)
    for attempt in range(MAX_RETRIES):
        if stop_event and stop_event.is_set():
            logging.info("Operação interrompida pelo usuário durante retry.")
            raise InterruptedError("Processo interrompido pelo usuário.")
        try:
            return func(*args, **kwargs)
        except Exception as e:
            if attempt == MAX_RETRIES - 1:
                raise e
            logging.warning(
                f"Tentativa {attempt + 1} falhou. Tentando novamente em {RETRY_DELAY} segundos: {e}"
            )
            time.sleep(RETRY_DELAY)

def get_month_weeks(year, month):
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, 1) + relativedelta(months=1) - timedelta(days=1)
    weeks = set()
    current_date = first_day
    while current_date <= last_day:
        weeks.add(current_date.isocalendar()[1])
        current_date += timedelta(days=1)
    return sorted(weeks)

def parse_interval(interval_str, current_year, current_month):
    try:
        if not interval_str:
            return None, None
        clean_str = re.sub(r"DIA\s*", "", str(interval_str).upper()).strip()
        parts = re.split(r"\s*-\s*", clean_str)
        if len(parts) != 2:
            return None, None
        start_day = int(parts[0])
        end_day = int(parts[1])
        if start_day > end_day:
            prev_month = current_month - 1
            prev_year = current_year
            if prev_month == 0:
                prev_month = 12
                prev_year -= 1
            start_date = datetime(prev_year, prev_month, start_day)
            end_date = datetime(current_year, current_month, end_day)
        else:
            start_date = datetime(current_year, current_month, start_day)
            end_date = datetime(current_year, current_month, end_day)
        return start_date.date(), end_date.date()
    except (ValueError, TypeError) as e:
        logging.error(f"Erro ao parsear intervalo '{interval_str}': {e}")
        return None, None

def buscar_arquivos_e_acessar_pastas(caminho_pasta, stop_event=None):
    if stop_event and stop_event.is_set():
        logging.info(f"Varredura interrompida na pasta: {caminho_pasta}")
        return []
    with cache_lock:
        if caminho_pasta in file_cache:
            return file_cache[caminho_pasta]
    arquivos = []
    try:
        if not retry(os.path.exists, caminho_pasta, stop_event=stop_event):
            logging.warning(f"Pasta não encontrada: {caminho_pasta}")
            return []
        for root, dirs, files in os.walk(caminho_pasta, topdown=True):
            if stop_event and stop_event.is_set():
                logging.info(f"Varredura interrompida na subpasta: {root}")
                return []
            for file in files:
                if os.path.splitext(file)[1].lower() in BACKUP_EXT:
                    full_path = os.path.join(root, file)
                    arquivos.append(full_path)
        with cache_lock:
            file_cache[caminho_pasta] = arquivos
        return arquivos
    except FileNotFoundError:
        logging.error(f"A pasta '{caminho_pasta}' não foi encontrada.")
        return []
    except PermissionError:
        logging.error(f"Permissão negada para acessar a pasta '{caminho_pasta}'.")
        return []
    except Exception as e:
        logging.error(f"Erro ao acessar a pasta {caminho_pasta}: {e}")
        return []

def buscar_arquivos_em_paralelo(caminhos_pastas, stop_event=None):
    resultados = {}
    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        future_to_path = {
            executor.submit(buscar_arquivos_e_acessar_pastas, caminho, stop_event): caminho
            for caminho, _ in caminhos_pastas
        }
        for future in as_completed(future_to_path):
            if stop_event and stop_event.is_set():
                logging.info("Busca paralela interrompida.")
                executor.shutdown(wait=False)
                return resultados
            caminho = future_to_path[future]
            try:
                resultados[caminho] = future.result(timeout=NETWORK_TIMEOUT)
            except Exception as e:
                logging.error(f"Erro ao processar pasta {caminho}: {e}")
                resultados[caminho] = []
    return resultados

def caminho_verde(dados_adicionados, arquivo_nome, arquivo_caminho, celula):
    celula.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    celula.value = dados_adicionados
    celula.font = Font(color="000000", bold=False)
    celula.alignment = Alignment(horizontal="center", vertical="center")

def caminho_vermelho(celula):
    celula.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    celula.value = "NOT FOUND"
    celula.font = Font(color="FFFFFF", bold=True)
    celula.alignment = Alignment(horizontal="center", vertical="center")

def encontrar_coluna_semana(ws, semana):
    for col in range(6, ws.max_column + 1):
        header_value = ws.cell(row=2, column=col).value
        if header_value is not None:
            match = re.search(r"(?:Semana\s*|\s*)(\d+)(?:\s*ª|\s*)", str(header_value), re.IGNORECASE)
            if match:
                semana_num = int(match.group(1))
                if semana_num == semana:
                    return col
    return None

def enviar_email_notificacao(wb, stop_event=None):
    if stop_event and stop_event.is_set():
        logging.info("Envio de e-mail interrompido pelo usuário.")
        return False
    missing_backups = defaultdict(lambda: {"semanas": [], "intervalos": []})
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        intervalos_por_semana = {}
        for col in range(6, ws.max_column + 1):
            if stop_event and stop_event.is_set():
                logging.info(f"Verificação de intervalos interrompida na aba: {sheet_name}")
                return False
            header_value = ws.cell(row=2, column=col).value
            if header_value is not None:
                match = re.search(r"(?:Semana\s*|\s*)(\d+)(?:\s*ª|\s*)", str(header_value), re.IGNORECASE)
                if match:
                    semana = int(match.group(1))
                    interval_str = ws.cell(row=3, column=col).value
                    start_date, end_date = parse_interval(interval_str, datetime.now().year, datetime.now().month)
                    if start_date and end_date:
                        intervalos_por_semana[semana] = (start_date, end_date)
        for col in range(6, ws.max_column + 1):
            if stop_event and stop_event.is_set():
                logging.info(f"Verificação de backups ausentes interrompida na aba: {sheet_name}")
                return False
            header_value = ws.cell(row=2, column=col).value
            if header_value is not None:
                match = re.search(r"(?:Semana\s*|\s*)(\d+)(?:\s*ª|\s*)", str(header_value), re.IGNORECASE)
                if match:
                    semana = int(match.group(1))
                    for row in range(4, ws.max_row + 1):
                        if stop_event and stop_event.is_set():
                            logging.info(f"Verificação de backups ausentes interrompida na aba: {sheet_name}")
                            return False
                        cell = ws.cell(row=row, column=col)
                        if cell.value == "NOT FOUND":
                            tag = ws.cell(row=row, column=1).value
                            responsible = ws.cell(row=row, column=2).value or "Não especificado"
                            setor = ws.cell(row=row, column=3).value
                            if tag and setor:
                                key = (sheet_name, tag, responsible, setor)
                                if semana in intervalos_por_semana:
                                    start_date, end_date = intervalos_por_semana[semana]
                                    missing_backups[key]["semanas"].append(semana)
                                    missing_backups[key]["intervalos"].append(
                                        f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
                                    )
                                    missing_backups[key].update({
                                        "sheet": sheet_name,
                                        "tag": tag,
                                        "responsible": responsible,
                                        "setor": setor,
                                    })
    if stop_event and stop_event.is_set():
        logging.info("Email interrompido pelo usuário antes de construir o corpo do e-mail.")
        return False
    if not missing_backups:
        logging.info("Nenhum backup ausente encontrado. Nenhum email enviado.")
        return False
    html_body = """
<html>
  <body>
    <h2 style="color: red;">Backups Ausentes Detectados</h2>
"""
    for key, miss in missing_backups.items():
        if not miss["semanas"]:
            continue
        semanas_str = ", ".join(map(str, sorted(miss["semanas"])))
        intervalos_str = ", ".join(miss["intervalos"])
        quantidade = len(miss["semanas"])
        html_body += f"""
        <div style="margin-bottom: 15px; padding: 10px; border: 1px solid #ccc; border-radius: 5px;">
            <p>TAG: <strong>{miss['tag']}</strong></p>
            <p>Responsável: <strong><span style="color: red;">{miss['responsible']}</span></strong></p>
            <p>Setor: <strong>{miss['setor']}</strong></p>
            <p>Semanas: {semanas_str}</p>
            <p>Quantidade de backups faltantes: <strong><span style="color: red;">{quantidade}</span></strong></p>
            <p>Intervalos: {intervalos_str}</p>
        </div>
    """
    html_body += """
  </body>
</html>
"""
    if stop_event and stop_event.is_set():
        logging.info("Email interrompido pelo usuário antes do envio.")
        return False
    msg = MIMEMultipart()
    msg["From"] = from_email
    msg["To"] = to_email if isinstance(to_email, str) else ", ".join(to_email)
    msg["Subject"] = "Relatório de Backups Ausentes"
    msg.attach(MIMEText(html_body, "html"))
    try:
        with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
            if stop_event and stop_event.is_set():
                logging.info("Conexão SMTP interrompida pelo usuário.")
                return False
            server.starttls()
            if stop_event and stop_event.is_set():
                logging.info("Autenticação SMTP interrompida pelo usuário.")
                return False
            server.login(smtp_user, smtp_pass)
            if stop_event and stop_event.is_set():
                logging.info("Envio de e-mail interrompido pelo usuário antes da entrega.")
                return False
            server.sendmail(from_email, to_email if isinstance(to_email, str) else to_email, msg.as_string())
        logging.info("E-mail ok")
        return True
    except smtplib.SMTPServerDisconnected:
        logging.error("Erro: Servidor SMTP desconectado inesperadamente. Verifique o servidor e a porta.")
        return False
    except smtplib.SMTPAuthenticationError:
        logging.error("Erro: Falha na autenticação. Verifique usuário e senha.")
        return False
    except Exception as e:
        logging.error(f"Erro ao enviar email: {e}")
        return False

def main(excel_path=None, send_email=True, stop_event=None, progress_callback=None):
    EXCEL_PATH = excel_path or EXCEL_PATH
    if stop_event and stop_event.is_set():
        logging.info("Verificação interrompida pelo usuário.")
        return False
    
    # Adicionar depuração para verificar STORAGE_BASE
    logging.info(f"Verificando caminho remoto: {STORAGE_BASE}")
    try:
        if os.path.exists(STORAGE_BASE):
            logging.info("Caminho remoto acessível.")
            dir_contents = os.listdir(STORAGE_BASE)
            logging.info(f"Conteúdo do diretório {STORAGE_BASE}: {dir_contents}")
        else:
            logging.error(f"Caminho remoto {STORAGE_BASE} não existe ou não está acessível.")
            # Tentar listar o diretório pai para depuração
            parent_path = os.path.dirname(STORAGE_BASE) 
            if os.path.exists(parent_path):
                logging.info(f"Conteúdo do diretório pai {parent_path}: {os.listdir(parent_path)}")
            else:
                logging.error(f"Diretório pai {parent_path} também não acessível.")
            raise Exception(f"Storage remoto {STORAGE_BASE} não acessível.")
    except Exception as e:
        logging.error(f"Erro ao acessar {STORAGE_BASE}: {e}", exc_info=True)
        return False

    SEMANA_ATUAL = datetime.now().isocalendar()[1]
    MES_ATUAL = datetime.now().month
    ANO_ATUAL = datetime.now().year
    if not os.path.exists(EXCEL_PATH):
        logging.error(f"Erro: O arquivo {EXCEL_PATH} não foi encontrado.")
        return False
    if not retry(os.path.exists, STORAGE_BASE, stop_event=stop_event):
        logging.error(f"Erro: Storage remoto {STORAGE_BASE} não acessível.")
        return False
    try:
        wb = retry(openpyxl.load_workbook, EXCEL_PATH, stop_event=stop_event)
        logging.info(f"Ano atual: {ANO_ATUAL}, Mês atual: {MES_ATUAL}, Semana atual: {SEMANA_ATUAL}")
        semanas_do_mes = get_month_weeks(ANO_ATUAL, MES_ATUAL)
        logging.info(f"Semanas do mês {MES_ATUAL}/{ANO_ATUAL}: {semanas_do_mes}")
        semanas_a_verificar = [semana for semana in semanas_do_mes if semana <= SEMANA_ATUAL]
        logging.info(f"Semanas a verificar: {semanas_a_verificar}")

        # Contar abas para progresso
        total_sheets = len(wb.sheetnames)
        current_sheet = 0

        for sheet_name in wb.sheetnames:
            if stop_event and stop_event.is_set():
                logging.info(f"Verificação interrompida na aba: {sheet_name}")
                return False
            ws = wb[sheet_name]
            intervalos_por_semana = {}
            meses_envolvidos = set()
            for col in range(6, ws.max_column + 1):
                if stop_event and stop_event.is_set():
                    logging.info(f"Verificação interrompida ao processar colunas da aba: {sheet_name}")
                    return False
                header_value = ws.cell(row=2, column=col).value
                if header_value is not None:
                    match = re.search(r"(?:Semana\s*|\s*)(\d+)(?:\s*ª|\s*)", str(header_value), re.IGNORECASE)
                    if match:
                        semana_num = int(match.group(1))
                        if semana_num in semanas_a_verificar:
                            interval_str = ws.cell(row=3, column=col).value
                            start_date, end_date = parse_interval(interval_str, ANO_ATUAL, MES_ATUAL)
                            if start_date and end_date:
                                intervalos_por_semana[semana_num] = (start_date, end_date)
                                meses_envolvidos.add(start_date.strftime("%m-%Y"))
                                meses_envolvidos.add(end_date.strftime("%m-%Y"))
            todas_as_pastas = []
            pasta_para_mes = {}
            for row in range(4, ws.max_row + 1):
                if stop_event and stop_event.is_set():
                    logging.info(f"Verificação interrompida ao processar linhas da aba: {sheet_name}")
                    return False
                informacao_coluna_A = ws.cell(row=row, column=1).value
                informacao_coluna_C = ws.cell(row=row, column=3).value
                if not informacao_coluna_A or not informacao_coluna_C:
                    continue
                for mes_ano in meses_envolvidos:
                    mes, ano = mes_ano.split("-")
                    caminho_formato = f"{informacao_coluna_C}/{informacao_coluna_A}/{mes_ano}"
                    caminho_pasta = os.path.join(STORAGE_BASE, ano, *caminho_formato.split("/"))
                    if caminho_pasta not in todas_as_pastas:
                        todas_as_pastas.append(caminho_pasta)
                        pasta_para_mes[caminho_pasta] = mes_ano
            resultados_busca = buscar_arquivos_em_paralelo([(pasta, pasta_para_mes[pasta]) for pasta in todas_as_pastas], stop_event=stop_event)
            for row in range(4, ws.max_row + 1):
                if stop_event and stop_event.is_set():
                    logging.info(f"Verificação interrompida ao processar linhas da aba: {sheet_name}")
                    return False
                informacao_coluna_A = ws.cell(row=row, column=1).value
                informacao_coluna_C = ws.cell(row=row, column=3).value
                if not informacao_coluna_A or not informacao_coluna_C:
                    continue
                arquivos_por_semana = defaultdict(list)
                for mes_ano in meses_envolvidos:
                    mes, ano = mes_ano.split("-")
                    caminho_formato = f"{informacao_coluna_C}/{informacao_coluna_A}/{mes_ano}"
                    caminho_pasta = os.path.join(STORAGE_BASE, ano, *caminho_formato.split("/"))
                    arquivos = resultados_busca.get(caminho_pasta, [])
                    for arquivo in arquivos:
                        try:
                            data_mod = datetime.fromtimestamp(os.path.getmtime(arquivo)).date()
                            ano_arquivo = data_mod.year
                            if ano_arquivo != ANO_ATUAL:
                                continue
                            for semana, (data_inicio, data_fim) in intervalos_por_semana.items():
                                if data_inicio <= data_mod <= data_fim:
                                    arquivos_por_semana[semana].append((arquivo, data_mod))
                        except OSError as e:
                            logging.error(f"Erro ao acessar arquivo {arquivo}: {e}")
                for semana in semanas_a_verificar:
                    if stop_event and stop_event.is_set():
                        logging.info(f"Verificação interrompida ao processar semana {semana} na aba: {sheet_name}")
                        return False
                    if semana not in intervalos_por_semana:
                        continue
                    coluna = encontrar_coluna_semana(ws, semana)
                    if coluna is None:
                        continue
                    celula = ws.cell(row=row, column=coluna)
                    if celula.value is not None and celula.value != "NOT FOUND":
                        continue
                    archs = arquivos_por_semana.get(semana, [])
                    if not archs:
                        caminho_vermelho(celula)
                        continue
                    archs.sort(key=lambda x: x[1], reverse=True)
                    arquivo_mais_recente, data_mod = archs[0]
                    data_str = data_mod.strftime(DATA_FORMATO)
                    caminho_verde(data_str, os.path.basename(arquivo_mais_recente), os.path.dirname(arquivo_mais_recente), celula)
            
            # Atualizar progresso após cada aba
            current_sheet += 1
            if progress_callback:
                progress_callback(current_sheet / total_sheets)

        retry(wb.save, EXCEL_PATH, stop_event=stop_event)
        logging.info("Dados gravados na planilha com sucesso.")
        
        # Forçar 100% ao final
        if progress_callback:
            progress_callback(1.0)
        
        if send_email:
            if enviar_email_notificacao(wb, stop_event=stop_event):
                logging.info("E-mail de notificação enviado com sucesso.")
                return True
            else:
                logging.info("Processo interrompido ou falha no envio de e-mail.")
                return False
        else:
            logging.info("Envio de e-mail desativado pelo usuário.")
            return True
    except Exception as e:
        logging.error(f"Erro geral: {e}", exc_info=True)
        return False
    finally:
        if "wb" in locals() and wb is not None:
            try:
                wb.close()
            except Exception as e:
                logging.error(f"Erro ao fechar o workbook: {e}")